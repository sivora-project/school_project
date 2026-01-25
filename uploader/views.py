from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Sum, F
from decimal import Decimal
import uuid
import openpyxl

from .models import (
    UserCredential, Student, Classes,
    StudentFee, FeePayment, FeeHead, Attendance
)
import json
from decimal import Decimal

from .utils import (
    get_student_attendance_summary,
    get_longest_attendance_streak,
    get_student_marks_summary
)
from collections import defaultdict
from django.db.models import Count
from .models import Student, Attendance, Classes
from django.contrib.auth.decorators import login_required
from datetime import date
from .models import Student, Attendance
from django.shortcuts import render, get_object_or_404
from django.db.models import Sum, F
from .models import Student, StudentFee
# ======================================================
# AUTH
# ======================================================
def user_login(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        user_rec = UserCredential.objects.filter(username__iexact=username).first()
        if not user_rec or user_rec.password != password:
            messages.error(request, "Invalid credentials")
            return render(request, "login_page.html")

        user, _ = User.objects.get_or_create(username=user_rec.username)
        login(request, user)

        request.session["role"] = (user_rec.role or "").upper()
        request.session["username"] = user_rec.username
        return redirect("dashboard")

    return render(request, "login_page.html")


def user_logout(request):
    logout(request)
    request.session.flush()
    return redirect("login")


@login_required
def dashboard(request):
    return render(request, "dashboard.html")


# ======================================================
# STUDENT INFO
# ======================================================
@login_required
def student_info(request):
    view_type = request.GET.get("view_type")
    context = {}

    if view_type == "individual":
        pen = request.GET.get("student_id")
        name = request.GET.get("student_name")
        month = request.GET.get("month")
        exam_code = request.GET.get("exam_code", "FA1")

        qs = Student.objects.all()
        if pen:
            qs = qs.filter(student_pen__iexact=pen)
        if name:
            qs = qs.filter(name__icontains=name)

        student = qs.first()

        if student:
            total, present, absent = get_student_attendance_summary(
                student.student_pen, month
            )
            streak = get_longest_attendance_streak(student.student_pen)
            marks_summary, subject_marks = get_student_marks_summary(
                student.student_id, exam_code
            )

            context.update({
                "show_dashboard": True,
                "student": student,
                "total_working_days": total,
                "present_days": present,
                "absent_days": absent,
                "longest_streak": streak,
                "subject_marks": subject_marks,
                "total_marks": marks_summary.get("total_marks", 0),
                "marks_obtained": marks_summary.get("obtained_marks", 0),
                "selected_exam": exam_code,
            })
        else:
            context["no_data"] = True

    return render(request, "student_info.html", context)

def attendance_for_students(attendance_qs, students_qs):
    """
    Safe helper to map Attendance → Students using student_pen
    """
    pens = list(students_qs.values_list("student_pen", flat=True))
    return attendance_qs.filter(student_pen__in=pens)


# ======================================================
# EXPORT SCHOOL DASHBOARD
# ======================================================
@login_required
def export_school_dashboard(request):
    snapshot_month = request.GET.get("snapshot_month")
    today = date.today()
    month = int(snapshot_month) if snapshot_month else today.month

    attendance_qs = Attendance.objects.filter(attendance_date__month=month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "School Attendance"

    ws.append(["Class", "Total Students", "Present", "Absent", "Attendance %"])

    for cls in Classes.objects.filter(is_active=True):
        students = Student.objects.filter(class_ref=cls)
        total = students.count()

        class_attendance = attendance_for_students(attendance_qs, students)
        present = class_attendance.filter(status="P").count()
        absent = total - present
        percent = round((present / total) * 100, 2) if total else 0

        ws.append([
            cls.class_name,
            total,
            present,
            absent,
            percent
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f"attachment; filename=school_dashboard_{month}.xlsx"
    )
    wb.save(response)
    return response




# ======================================================
# STAFF FEE DASHBOARD (DISPLAY ONLY)
# ======================================================
def fee_dashboard(request):
    student = None
    fee_rows = []
    selected_fee = None
    error = None

    # --------------------
    # SEARCH BY PEN
    # --------------------
    if request.method == "POST" and "student_pen" in request.POST:
        pen = request.POST.get("student_pen")

        student = Student.objects.filter(student_pen=pen).first()
        if not student:
            return render(request, "staff_fee_dashboard.html", {
                "error": "Invalid Student PEN"
            })

        fee_rows = StudentFee.objects.filter(
            student_id=student.student_id
        ).select_related("fee_head")

    # --------------------
    # PAY NOW CLICK
    # --------------------
    if request.method == "GET" and request.GET.get("pay"):
        fee_id = request.GET.get("pay")

        selected_fee = StudentFee.objects.filter(id=fee_id).first()
        if selected_fee:
            student = Student.objects.filter(
                student_id=selected_fee.student_id
            ).first()

            fee_rows = StudentFee.objects.filter(
                student_id=student.student_id
            ).select_related("fee_head")


    # --------------------
    # PAYMENT SUBMIT
    # --------------------
    if request.method == "POST" and "amount" in request.POST:
        fee_id = request.POST.get("student_fee_id")
        amount = Decimal(request.POST.get("amount"))
        mode = request.POST.get("payment_mode")
        utr = request.POST.get("utr_number")

        selected_fee = StudentFee.objects.filter(id=fee_id).first()
        student = Student.objects.filter(
            student_id=selected_fee.student_id
        ).first()

        if amount > selected_fee.due_amount:
            error = "Amount exceeds balance"

        else:
            with transaction.atomic():
                receipt_no = generate_receipt_no()

                payment = FeePayment.objects.create(
                    student_fee=selected_fee,
                    payment_date=date.today(),
                    amount=amount,
                    payment_mode=mode,
                    receipt_no=receipt_no,
                    utr_number=utr if mode == "Non-Cash" else None
                )

                from django.db.models import F

                StudentFee.objects.filter(
                    id=selected_fee.id
                ).update(
                    paid_amount=F("paid_amount") + amount
                )

                # selected_fee.paid_amount += amount
                # selected_fee.save()


            return redirect("fee_receipt", payment_id=payment.id)

    return render(request, "staff_fee_dashboard.html", {
        "student": student,
        "fee_rows": fee_rows,
        "selected_fee": selected_fee,
        "error": error
    })
# ======================================================
# PAY FEE (SEPARATE FUNCTION – FINAL)
# ======================================================
@login_required
def pay_fee(request, student_fee_id):
    student_fee = StudentFee.objects.filter(id=student_fee_id).first()
    if not student_fee:
        return HttpResponse("Fee record not found", status=404)

    student = Student.objects.filter(
        student_id=student_fee.student_id
    ).first()

    if not student:
        return HttpResponse("Student not found", status=404)

    if request.method == "POST":
        amount = float(request.POST.get("amount"))
        mode = request.POST.get("payment_mode")
        utr = request.POST.get("utr")

        if amount > student_fee.due_amount:
            return render(request, "pay_fee.html", {
                "error": "Amount exceeds balance",
                "student_fee": student_fee,
                "student": student
            })

        with transaction.atomic():
            payment = FeePayment.objects.create(
                student_fee=student_fee,
                payment_date=date.today(),
                amount=amount,
                payment_mode=mode,
                receipt_no=f"RCPT-{date.today().year}-{student_fee_id}"
            )

            student_fee.paid_amount += amount
            student_fee.save()

        return redirect("fee_receipt", payment.id)

    return render(request, "staff_fee_dashboard.html", {
        "student_fee": student_fee,
        "student": student
    })


# ======================================================
# RECEIPT
# ======================================================
@login_required
def fee_receipt(request, payment_id):
    payment = get_object_or_404(FeePayment, id=payment_id)
    fee = payment.student_fee
    student = get_object_or_404(Student, student_id=fee.student_id)

    return render(request, "fee_receipt.html", {
        "payment": payment,
        "student": student,
        "fee": fee
    })


def generate_receipt_no():
    today = date.today().strftime("%Y%m%d")
    return f"RCPT-{today}-{uuid.uuid4().hex[:6].upper()}"


# ======================================================
# DAILY COLLECTION REPORT
# ======================================================
@login_required
def daily_collection_report(request):
    selected_date = request.GET.get("date")

    if selected_date:
        report_date = selected_date
    else:
        report_date = date.today()

    payments = FeePayment.objects.filter(
        payment_date=report_date
    ).select_related("student_fee")

    total_amount = payments.aggregate(
        total=Sum("amount")
    )["total"] or 0

    cash_amount = payments.filter(
        payment_mode="Cash"
    ).aggregate(total=Sum("amount"))["total"] or 0

    non_cash_amount = payments.filter(
        payment_mode="Non-Cash"
    ).aggregate(total=Sum("amount"))["total"] or 0

    report_rows = []
    for p in payments:
        student = Student.objects.filter(
            student_id=p.student_fee.student_id
        ).first()

        report_rows.append({
            "receipt_no": p.receipt_no,
            "student_name": student.name if student else "",
            "fee_type": (
                f"Term {p.student_fee.term_no} Fee"
                if p.student_fee.fee_head.is_term_fee
                else p.student_fee.fee_head.fee_name
            ),
            "amount": p.amount,
            "payment_mode": p.payment_mode,
            "utr": getattr(p, "utr_number", "")
        })

    return render(request, "daily_collection_report.html", {
        "report_date": report_date,
        "total_amount": total_amount,
        "cash_amount": cash_amount,
        "non_cash_amount": non_cash_amount,
        "report_rows": report_rows
    })


# ======================================================
# AUTHORITY DASHBOARD
# ======================================================
@login_required
def authority_fee_dashboard(request):
    return render(request, "authority_fee_dashboard.html", {
        "classes": Classes.objects.all(),
        "fee_heads": FeeHead.objects.all()
    })


@login_required
def authority_fee_chart_data(request):
    academic_year = request.GET.get("academic_year")
    class_id = request.GET.get("class_id")
    fee_head_id = request.GET.get("fee_head_id")

    fees = StudentFee.objects.select_related("fee_head")

    if academic_year:
        fees = fees.filter(academic_year=academic_year)
    if fee_head_id:
        fees = fees.filter(fee_head_id=fee_head_id)
    if class_id:
        student_ids = Student.objects.filter(
            class_ref_id=class_id
        ).values_list("student_id", flat=True)
        fees = fees.filter(student_id__in=student_ids)

    total = fees.aggregate(t=Sum("total_amount"))["t"] or 0
    collected = fees.aggregate(t=Sum("paid_amount"))["t"] or 0

    fee_head = FeeHead.objects.filter(id=fee_head_id).first()

    if fee_head and fee_head.is_term_fee:
        data = fees.exclude(term_no__isnull=True).values("term_no").annotate(
            collected=Sum("paid_amount")
        ).order_by("term_no")
        labels = [f"Term {d['term_no']}" for d in data]
        values = [float(d["collected"] or 0) for d in data]
    else:
        chart = {}
        for f in fees:
            stu = Student.objects.filter(student_id=f.student_id).first()
            cls_name = stu.class_ref.class_name if stu and stu.class_ref else "UNKNOWN"
            chart[cls_name] = chart.get(cls_name, 0) + float(f.paid_amount)
        labels = list(chart.keys())
        values = list(chart.values())

    return JsonResponse({
        "kpi": {
            "total": total,
            "collected": collected,
            "pending": total - collected,
            "percentage": round((collected / total) * 100, 2) if total else 0
        },
        "labels": labels,
        "values": values
    })


# ======================================================
# DRILL – CLASS → STUDENTS
# ======================================================
@login_required
def authority_fee_students(request):
    class_id = request.GET.get("class_id")
    students = Student.objects.filter(class_ref_id=class_id)

    data = []
    for s in students:
        fees = StudentFee.objects.filter(student_id=s.student_id)
        total = fees.aggregate(t=Sum("total_amount"))["t"] or 0
        paid = fees.aggregate(t=Sum("paid_amount"))["t"] or 0

        data.append({
            "student_id": s.student_id,
            "pen": s.student_pen,
            "name": s.name,
            "total": total,
            "paid": paid,
            "pending": total - paid
        })

    return JsonResponse(data, safe=False)


# ======================================================
# DRILL – STUDENT PROFILE (READ ONLY)
# ======================================================
@login_required
def authority_student_profile(request, student_id):
    student = get_object_or_404(Student, student_id=student_id)
    fees = StudentFee.objects.filter(student_id=student.student_id)

    return render(request, "authority_student_profile.html", {
        "student": student,
        "fees": fees
    })


# ======================================================
# EXPORT AUTHORITY DASHBOARD
# ======================================================
@login_required
def authority_export_excel(request):
    class_id = request.GET.get("class_id")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Student PEN", "Name", "Total Fee", "Paid", "Pending"])

    students = Student.objects.filter(class_ref_id=class_id)
    for s in students:
        fees = StudentFee.objects.filter(student_id=s.student_id)
        total = fees.aggregate(t=Sum("total_amount"))["t"] or 0
        paid = fees.aggregate(t=Sum("paid_amount"))["t"] or 0

        ws.append([s.student_pen, s.name, total, paid, total - paid])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=authority_fee_report.xlsx"
    wb.save(response)
    return response




# =========================================================
# REGULAR ABSENT DRILL
# =========================================================
@login_required
def regular_absent_drill(request):

    class_name = request.GET.get("class_name")
    snapshot_month = request.GET.get("snapshot_month")


    year, month = map(int, snapshot_month.split("-"))

    students = Student.objects.filter(
        class_ref__class_name=class_name
    )

    student_pens = students.values_list("student_pen", flat=True)

    # Count absent days per student
    absent_counts = (
        Attendance.objects
        .filter(
            student_pen__in=student_pens,
            status="A",
            attendance_date__year=year,
            attendance_date__month=month
        )
        .values("student_pen")
        .annotate(absent_days=Count("id"))
        .filter(absent_days__gte=3)
    )

    # Map for quick lookup
    absent_map = {
        a["student_pen"]: a["absent_days"]
        for a in absent_counts
    }

    rows = []
    for s in students:
        if s.student_pen in absent_map:
            rows.append({
                "student_pen": s.student_pen,
                "name": s.name,
                "absent_days": absent_map[s.student_pen],
            })

    return render(request, "regular_absent_drill.html", {
        "class_name": class_name,
        "snapshot_month": snapshot_month,
        "rows": rows,
    })

# =========================================================
# SCHOOL DASHBOARD
# =========================================================


@login_required
def school_dashboard(request):
    today = date.today()

    snapshot_mode = request.GET.get("snapshot_mode", "month")
    snapshot_month = request.GET.get("snapshot_month")
    snapshot_date = request.GET.get("snapshot_date")
    selected_class = request.GET.get("class_name", "ALL")

    # -------------------------
    # Month / Date filter
    # -------------------------
    if snapshot_month:
        year, month = map(int, snapshot_month.split("-"))
    else:
        year, month = today.year, today.month
        snapshot_month = f"{year}-{month:02d}"

    # -------------------------
    # Base student list (only valid classes)
    # -------------------------
    students = Student.objects.select_related("class_ref").exclude(class_ref__isnull=True)

    if selected_class != "ALL":
        students = students.filter(class_ref__class_name=selected_class)

    total_students = students.count()
    student_pens = list(students.values_list("student_pen", flat=True))

    # -------------------------
    # Attendance data (filtered)
    # -------------------------
    attendance_qs = Attendance.objects.filter(
        student_pen__in=student_pens,
        attendance_date__year=year,
        attendance_date__month=month
    )

    if snapshot_mode == "day" and snapshot_date:
        attendance_qs = attendance_qs.filter(attendance_date=snapshot_date)
        snapshot_label = f"Date: {snapshot_date}"
    else:
        snapshot_label = f"Month: {snapshot_month}"

    # -------------------------
    # KPI: Present / Absent
    # -------------------------
    present_students = attendance_qs.filter(status="P").values("student_pen").distinct().count()
    absent_students = attendance_qs.filter(status="A").values("student_pen").distinct().count()

    attendance_percentage = round(
        (present_students / total_students) * 100, 2
    ) if total_students else 0

    # -------------------------
    # Class-wise merge (A+B → one row)
    # -------------------------
    class_map = defaultdict(lambda: {"sections": set(), "students": set()})

    for s in students:
        cls = s.class_ref
        if not cls:
            continue
        class_map[cls.class_name]["sections"].add(cls.section)
        class_map[cls.class_name]["students"].add(s.student_pen)

    class_totals = []
    for class_name, data in class_map.items():
        pens = list(data["students"])
        total = len(pens)

        present = attendance_qs.filter(student_pen__in=pens, status="P").values("student_pen").distinct().count()
        absent = attendance_qs.filter(student_pen__in=pens, status="A").values("student_pen").distinct().count()

        percent = round((present / total) * 100, 2) if total else 0

        class_totals.append({
            "class_name": class_name,
            "sections": "+".join(sorted(data["sections"])),
            "total": total,
            "present": present,
            "absent": absent,
            "percent": percent,
        })

    # -------------------------
    # Regular absentees (>= 3 days)
    # -------------------------
    absent_counts = (
        attendance_qs.filter(status="A")
        .values("student_pen")
        .annotate(days=Count("id"))
        .filter(days__gte=3)
    )

    pen_absent_map = {a["student_pen"]: a["days"] for a in absent_counts}

    regular_absent_map = defaultdict(int)
    for s in students:
        if pen_absent_map.get(s.student_pen, 0) >= 3:
            if s.class_ref:
                regular_absent_map[s.class_ref.class_name] += 1

    regular_absentees = [
        {"class_name": k, "count": v}
        for k, v in regular_absent_map.items()
    ]

    # -------------------------
    # Day-wise trend
    # -------------------------
    trend_labels, trend_data = [], []
    if snapshot_mode == "month":
        for day in range(1, 32):
            try:
                d = date(year, month, day)
            except ValueError:
                continue

            cnt = attendance_qs.filter(attendance_date=d, status="P").count()
            trend_labels.append(str(day))
            trend_data.append(cnt)

    # -------------------------
    # Context
    # -------------------------
    return render(request, "school_dashboard.html", {
        "snapshot_label": snapshot_label,
        "snapshot_mode": snapshot_mode,
        "snapshot_month": snapshot_month,
        "snapshot_date": snapshot_date,
        "selected_class": selected_class,

        "total_students": total_students,
        "total_present": present_students,
        "total_absent": absent_students,
        "attendance_percentage": attendance_percentage,

        "class_totals": class_totals,
        "regular_absentees": regular_absentees,

        "trend_labels": trend_labels,
        "trend_data": trend_data,

        "all_classes": Classes.objects.values_list("class_name", flat=True).distinct(),
    })






@login_required
def class_drill_dashboard(request):
    class_name = request.GET.get("class_name")
    snapshot_mode = request.GET.get("snapshot_mode", "month")
    snapshot_month = request.GET.get("snapshot_month")
    snapshot_date = request.GET.get("snapshot_date")

    # -------------------------
    # Resolve date / month
    # -------------------------
    if snapshot_mode == "day" and snapshot_date:
        year, month, day = map(int, snapshot_date.split("-"))
    else:
        year, month = map(int, snapshot_month.split("-"))
        day = None

    # -------------------------
    # Students in class
    # -------------------------
    students = Student.objects.filter(
        class_ref__class_name=class_name
    ).select_related("class_ref")

    # -------------------------
    # Attendance for period
    # -------------------------
    attendance_qs = Attendance.objects.filter(
        student_pen__in=students.values_list("student_pen", flat=True),
        attendance_date__year=year,
        attendance_date__month=month
    )

    if snapshot_mode == "day" and day:
        attendance_qs = attendance_qs.filter(attendance_date__day=day)

    # ⚠️ IMPORTANT FIX
    # One student can have ONLY ONE record per day
    attendance_map = {
        a.student_pen: a.status
        for a in attendance_qs
    }

    # -------------------------
    # Prepare rows
    # -------------------------
    rows = []
    sections = set()

    for s in students:
        sections.add(s.class_ref.section)
        rows.append({
            "student_pen": s.student_pen,
            "name": s.name,
            "status": attendance_map.get(s.student_pen, "A")  # default ABSENT
        })

    # -------------------------
    # Context
    # -------------------------
    context = {
        "class_name": class_name,
        "section": "+".join(sorted(sections)),
        "student_rows": rows,
        "snapshot_mode": snapshot_mode,
        "snapshot_month": snapshot_month,
        "snapshot_date": snapshot_date,
    }

    return render(request, "class_drill_dashboard.html", context)

from django.shortcuts import render, get_object_or_404
from django.db.models import Sum
from .models import Student, StudentFee


from django.shortcuts import render, get_object_or_404
from django.db.models import Sum
from .models import Student, StudentFee


from django.shortcuts import render
from django.db.models import Sum
from .models import Student, StudentFee, Classes, FeeHead

from django.shortcuts import render
from django.db.models import Sum
from decimal import Decimal
import json

from .models import Student, StudentFee, Classes, FeeHead


def management_fee_dashboard(request):
    class_filter = request.GET.get("class")
    section_filter = request.GET.get("section")
    fee_type_filter = request.GET.get("fee_type")
    search = request.GET.get("search")

    # -----------------------------
    # DROPDOWN MASTER DATA
    # -----------------------------
    class_list = Classes.objects.values("class_name").distinct().order_by("class_name")
    section_list = Student.objects.values_list("section", flat=True).distinct()
    fee_type_list = FeeHead.objects.all().order_by("fee_name")

    # -----------------------------
    # 1️⃣ FILTER STUDENTS FIRST
    # -----------------------------
    students_qs = Student.objects.select_related("class_ref").all()

    if class_filter:
        students_qs = students_qs.filter(class_ref_id=class_filter)

    if section_filter:
        students_qs = students_qs.filter(section=section_filter)

    if search:
        students_qs = students_qs.filter(
            name__icontains=search
        ) | students_qs.filter(student_pen__icontains=search)

    students_map = {s.student_id: s for s in students_qs}
    student_ids = list(students_map.keys())

    # -----------------------------
    # 2️⃣ FILTER FEES USING STUDENTS
    # -----------------------------
    fee_qs = StudentFee.objects.filter(student_id__in=student_ids)

    if fee_type_filter:
        fee_qs = fee_qs.filter(fee_head_id=fee_type_filter)

    # -----------------------------
    # 3️⃣ STUDENT-WISE AGGREGATION
    # -----------------------------
    fee_summary = (
        fee_qs
        .values("student_id")
        .annotate(
            total_expected=Sum("total_amount"),
            total_paid=Sum("paid_amount"),
            pending_amount=Sum("total_amount") - Sum("paid_amount"),
        )
        .filter(pending_amount__gt=0)
    )

    student_due_list = []

    for row in fee_summary:
        student = students_map.get(row["student_id"])
        if not student or not student.class_ref:
            continue

        student_due_list.append({
            "student_pen": student.student_pen,
            "name": student.name,
            "class_name": student.class_ref.class_name,  # FK ONLY
            "section": student.section,
            "total_expected": row["total_expected"],
            "total_paid": row["total_paid"],
            "pending_amount": row["pending_amount"],
        })

    # -----------------------------
    # 4️⃣ PIE CHART (CLASS-WISE)
    # -----------------------------
    pie_map = {}

    for row in fee_summary:
        student = students_map.get(row["student_id"])
        if not student or not student.class_ref:
            continue

        cls = student.class_ref.class_name
        pie_map[cls] = pie_map.get(cls, Decimal("0.0")) + row["pending_amount"]

    # -----------------------------
    # 5️⃣ KPI CARDS
    # -----------------------------
    kpis = fee_qs.aggregate(
        total_expected=Sum("total_amount"),
        total_paid=Sum("paid_amount"),
        total_pending=Sum("total_amount") - Sum("paid_amount"),
    )

    return render(
        request,
        "management_fee_dashboard.html",
        {
            "student_due_list": student_due_list,
            "kpis": kpis,
            "class_list": class_list,
            "section_list": section_list,
            "fee_type_list": fee_type_list,
            "class_filter": class_filter,
            "section_filter": section_filter,
            "fee_type_filter": fee_type_filter,
            "search": search,
            # JSON SAFE FOR CHART.JS
            "pie_labels": json.dumps(list(pie_map.keys())),
            "pie_data": json.dumps([float(v) for v in pie_map.values()]),
        },
    )

def management_student_fee_detail(request, student_pen):
    student = get_object_or_404(Student, student_pen=student_pen)

    fee_rows = StudentFee.objects.filter(student_id=student.student_id)

    summary = fee_rows.aggregate(
        total_expected=Sum("total_amount"),
        total_paid=Sum("paid_amount"),
        pending=Sum("total_amount") - Sum("paid_amount"),
    )

    return render(
        request,
        "management_student_fee_detail.html",
        {
            "student": student,
            "fee_rows": fee_rows,
            "summary": summary,
        },
    )



# ==================================
# EXPORT ATTENDANCE TO EXCEL (FIXED)
# ==================================
def export_attendance_excel(request):
    class_id = request.GET.get("class_id")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    ws.append([
        "Student Name",
        "Total Days",
        "Present Days",
        "Absent Days",
        "Attendance %"
    ])

    students = Student.objects.filter(class_ref_id=class_id)

    for student in students:
        total = Attendance.objects.filter(student_pen=student.student_pen).count()
        present = Attendance.objects.filter(
            student_pen=student.student_pen,
            status="Present"
        ).count()
        absent = Attendance.objects.filter(
            student_pen=student.student_pen,
            status="Absent"
        ).count()

        percent = round((present / total) * 100, 2) if total else 0

        ws.append([
            student.name,
            total,
            present,
            absent,
            percent
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=attendance_report.xlsx"
    wb.save(response)

    return response



from datetime import date
from django.shortcuts import render
from django.contrib import messages
from collections import defaultdict

from .models import (
    Student,
    Attendance,
    Classes,
    AcademicCalendar
)

from datetime import date
from django.shortcuts import render, redirect
from django.contrib import messages

from .models import Student, Attendance, Classes, AcademicCalendar


def mark_attendance(request):
    # DISTINCT active classes
    classes = (
        Classes.objects
        .filter(is_active=True)
        .order_by("class_name")
        .distinct("class_name")
    )

    students = []
    present_count = 0
    absent_count = 0

    # -------------------------
    # READ FROM GET (LOAD)
    # -------------------------
    selected_class = request.GET.get("class_id")
    attendance_date = request.GET.get("attendance_date")

    attendance_date = (
        date.fromisoformat(attendance_date)
        if attendance_date else date.today()
    )

    if attendance_date > date.today():
        attendance_date = date.today()
        messages.warning(request, "Future dates are not allowed.")

    # -------------------------
    # ACADEMIC CALENDAR
    # -------------------------
    calendar_entry = AcademicCalendar.objects.filter(date=attendance_date).first()
    is_working_day = calendar_entry.is_working_day if calendar_entry else False
    non_working_reason = calendar_entry.reason if calendar_entry else "Non-working day"

    # -------------------------
    # LOAD STUDENTS
    # -------------------------
    if selected_class:
        students = list(
            Student.objects.filter(class_ref_id=selected_class)
        )

        student_pens = [s.student_pen for s in students]

        daily_records = Attendance.objects.filter(
            student_pen__in=student_pens,
            attendance_date=attendance_date
        )

        daily_map = {r.student_pen: r for r in daily_records}

        present_count = daily_records.filter(status="P").count()
        absent_count = daily_records.filter(status="A").count()

        # ---- Monthly register days
        year = attendance_date.year
        month = attendance_date.month

        working_dates = Attendance.objects.filter(
            student_pen__in=student_pens,
            attendance_date__year=year,
            attendance_date__month=month
        ).values_list("attendance_date", flat=True).distinct()

        total_working_days = len(working_dates)

        # ---- Academic Year (Apr–Mar)
        if attendance_date.month >= 4:
            ay_start = date(attendance_date.year, 4, 1)
            ay_end = date(attendance_date.year + 1, 3, 31)
        else:
            ay_start = date(attendance_date.year - 1, 4, 1)
            ay_end = date(attendance_date.year, 3, 31)

        for student in students:
            record = daily_map.get(student.student_pen)

            student.saved_status = (
                "Present" if record and record.status == "P"
                else "Absent" if record else "Present"
            )

            student.remarks = record.remarks if record else ""

            # Monthly %
            monthly_present = Attendance.objects.filter(
                student_pen=student.student_pen,
                attendance_date__in=working_dates,
                status="P"
            ).count()

            student.monthly_percentage = round(
                (monthly_present / total_working_days) * 100, 2
            ) if total_working_days else 0

            # Academic %
            ay_records = Attendance.objects.filter(
                student_pen=student.student_pen,
                attendance_date__gte=ay_start,
                attendance_date__lte=ay_end
            )

            ay_present = ay_records.filter(status="P").count()
            ay_total = ay_records.count()

            student.yearly_percentage = round(
                (ay_present / ay_total) * 100, 2
            ) if ay_total else 0

    # -------------------------
    # SAVE ATTENDANCE (POST)
    # -------------------------
    if request.method == "POST":
        selected_class = request.POST.get("class_id")
        attendance_date = date.fromisoformat(request.POST.get("attendance_date"))

        if not is_working_day:
            messages.error(
                request,
                f"{attendance_date} is a non-working day ({non_working_reason})."
            )
            return redirect(
                request.path + f"?class_id={selected_class}&attendance_date={attendance_date}"
            )

        for student in students:
            status_ui = request.POST.get(f"status_{student.student_id}")
            remarks = request.POST.get(f"remarks_{student.student_id}", "").strip()

            if attendance_date < date.today():
                if len(remarks.replace(" ", "")) < 10:
                    messages.error(
                        request,
                        f"Remarks must be at least 10 characters for {student.name}"
                    )
                    continue

                updated_by = request.user.get_username()
                final_remarks = f"Updated by: {updated_by}, Reason: {remarks}"
            else:
                final_remarks = remarks if remarks else None

            db_status = "P" if status_ui == "Present" else "A"

            Attendance.objects.update_or_create(
                student_pen=student.student_pen,
                attendance_date=attendance_date,
                defaults={
                    "status": db_status,
                    "remarks": final_remarks
                }
            )

        messages.success(request, "Attendance saved successfully")

        return redirect(
            request.path + f"?class_id={selected_class}&attendance_date={attendance_date}"
        )

    return render(request, "mark_attendance.html", {
        "classes": classes,
        "students": students,
        "selected_class": selected_class,
        "attendance_date": attendance_date,
        "is_working_day": is_working_day,
        "non_working_reason": non_working_reason,
        "present_count": present_count,
        "absent_count": absent_count,
    })



##---- performance dashbaord (Marks )




from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db.models import Avg, Count
import openpyxl

from .models import StudentMarks, ExamMaster


# -------------------------------
# DASHBOARD PAGE
# -------------------------------
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db.models import Avg, Count, Q
import openpyxl

from uploader.models import StudentMarks, ExamMaster


# -------------------------------------------------
# DASHBOARD PAGE
# -------------------------------------------------
def class_performance_dashboard(request):
    exams = ExamMaster.objects.all().order_by("term_no", "exam_code")
    return render(
        request,
        "dashboard/class_performance_dashboard.html",
        {"exams": exams}
    )


# -------------------------------------------------
# CLASS PERFORMANCE (EXAM-WISE)
# -------------------------------------------------
def class_performance_api(request):
    exam_code = request.GET.get("exam_code")

    qs = (
        StudentMarks.objects
        .filter(exam_code=exam_code)
        .filter(Q(is_absent=False) | Q(is_absent__isnull=True))
        .select_related("class_ref")
        .values(
            "class_ref_id",
            "class_ref__class_name",
            "class_ref__section"
        )
        .annotate(
            avg_marks=Avg("marks_obtained"),
            students=Count("student", distinct=True)
        )
    )

    labels, bar_data, table = [], [], []
    total_students = total_passed = 0
    best_class, best_avg = "", 0
    excellent = average = poor = 0

    for r in qs:
        avg = round(r["avg_marks"], 2)
        cname = f'{r["class_ref__class_name"]}-{r["class_ref__section"]}'

        labels.append(cname)
        bar_data.append(avg)

        total_students += r["students"]
        if avg >= 35:
            total_passed += r["students"]

        if avg > best_avg:
            best_avg = avg
            best_class = cname

        if avg >= 75:
            excellent += 1
        elif avg >= 50:
            average += 1
        else:
            poor += 1

        table.append({
            "class_id": r["class_ref_id"],
            "class": cname,
            "students": r["students"],
            "avg": avg
        })

    return JsonResponse({
        "kpi": {
            "total_classes": len(table),
            "best_class": best_class,
            "pass_percentage": round(
                (total_passed / total_students) * 100, 2
            ) if total_students else 0
        },
        "bar": {"labels": labels, "data": bar_data},
        "doughnut": {
            "labels": ["Excellent", "Average", "Needs Improvement"],
            "data": [excellent, average, poor]
        },
        "table": table
    })


# -------------------------------------------------
# SUBJECT-WISE PERFORMANCE
# -------------------------------------------------
def subject_performance_api(request):
    exam_code = request.GET.get("exam_code")
    class_id = request.GET.get("class_id")

    qs = (
        StudentMarks.objects
        .filter(
            exam_code=exam_code,
            class_ref__id=class_id
        )
        .filter(Q(is_absent=False) | Q(is_absent__isnull=True))
        .values("subject_code")
        .annotate(avg_marks=Avg("marks_obtained"))
        .order_by("subject_code")
    )

    return JsonResponse({
        "labels": list(qs.values_list("subject_code", flat=True)),
        "data": [round(x, 2) for x in qs.values_list("avg_marks", flat=True)]
    })


# -------------------------------------------------
# STUDENT-WISE MARKS (PIVOT + TOTAL)
# -------------------------------------------------
def student_wise_marks_api(request):
    exam_code = request.GET.get("exam_code")
    class_id = request.GET.get("class_id")

    qs = (
        StudentMarks.objects
        .filter(
            exam_code=exam_code,
            class_ref__id=class_id
        )
        .filter(Q(is_absent=False) | Q(is_absent__isnull=True))
        .select_related("student")
        .values(
            "student__student_id",
            "student__name",
            "subject_code",
            "marks_obtained"
        )
        .order_by("student__name")
    )

    subjects = sorted({r["subject_code"] for r in qs})

    students = {}
    subject_totals = {s: 0 for s in subjects}
    grand_total = 0

    for r in qs:
        sid = r["student__student_id"]
        marks = r["marks_obtained"]

        if sid not in students:
            students[sid] = {
                "student": r["student__name"],
                "total": 0,
                **{s: "-" for s in subjects}
            }

        students[sid][r["subject_code"]] = marks
        students[sid]["total"] += marks

        subject_totals[r["subject_code"]] += marks
        grand_total += marks

    return JsonResponse({
        "subjects": subjects,
        "rows": list(students.values()),
        "subject_totals": subject_totals,
        "grand_total": grand_total
    })



# -------------------------------
# EXPORT EXCEL
# -------------------------------
def export_class_performance_excel(request):
    exam_code = request.GET.get("exam_code")

    qs = (
        StudentMarks.objects
        .filter(exam_code=exam_code)
        .exclude(is_absent=True)
        .select_related("class_ref")
        .values(
            "class_ref__class_name",
            "class_ref__section"
        )
        .annotate(
            avg_marks=Avg("marks_obtained"),
            students=Count("student", distinct=True)
        )
        .order_by("class_ref__class_name")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Class Performance"
    ws.append(["Class", "Students", "Average Marks"])

    for r in qs:
        ws.append([
            f'{r["class_ref__class_name"]}-{r["class_ref__section"]}',
            r["students"],
            round(r["avg_marks"], 2)
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="class_performance_{exam_code}.xlsx"'
    )
    wb.save(response)
    return response


from django.http import HttpResponse
from django.db.models import Avg, Count
import openpyxl

from .models import StudentMarks


# -------------------------------------------------
# EXPORT CLASS PERFORMANCE (EXAM-WISE)
# -------------------------------------------------
def export_class_performance_excel(request):
    exam_code = request.GET.get("exam_code")

    # Safety check
    if not exam_code:
        return HttpResponse("Exam code is required", status=400)

    qs = (
        StudentMarks.objects
        .filter(
            exam_code=exam_code,
            is_absent=False
        )
        .select_related("class_ref")
        .values(
            "class_ref__class_name",
            "class_ref__section"
        )
        .annotate(
            avg_marks=Avg("marks_obtained"),
            students=Count("student", distinct=True)
        )
        .order_by("class_ref__class_name", "class_ref__section")
    )

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Class Performance"

    # Header
    ws.append([
        "Class",
        "Students",
        "Average Marks"
    ])

    # Data rows
    for r in qs:
        class_name = f'{r["class_ref__class_name"]}-{r["class_ref__section"]}'
        ws.append([
            class_name,
            r["students"],
            round(r["avg_marks"], 2) if r["avg_marks"] is not None else 0
        ])

    # Response
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        f'attachment; filename="class_performance_{exam_code}.xlsx"'
    )

    wb.save(response)
    return response
